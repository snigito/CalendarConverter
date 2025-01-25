import openpyxl
from datetime import datetime, timedelta
from ics import Calendar, Event
import sys
import pytz

# Function to parse date and time from Excel format
def parse_datetime(date_str, time_str):
    date = datetime.strptime(date_str, "%m/%d/%Y")
    time = datetime.strptime(time_str, "%I:%M %p").time()
    return datetime.combine(date, time)


def parse_course_entries(entries):
    meetings = []

    entry_list = entries.split('\n')

    for entry in entry_list:
        entry = entry.strip()
        if not entry:
            continue

        parts = entry.split('|')
        if len(parts) < 3:
            print(f"Skipping malformed entry: {entry}")
            continue

        days_part = parts[0].strip()
        time_interval = parts[1].strip()
        location = parts[2].strip() if len(parts) > 2 else "TBD"

        days = days_part.split('/')

        for day in days:
            meetings.append((day.strip(), time_interval, location))

    return meetings

def get_week_dates(start_date, end_date, day_of_week):
    days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_idx = days_of_week.index(day_of_week)
    current_date = start_date + timedelta(days=(day_idx - start_date.weekday() + 7) % 7)
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(weeks=1)

# Function to read the Excel file and print event details
def process_class_info(file_path, ics_path):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    est = pytz.timezone('US/Eastern')
    cal = Calendar()

    # Find the column indices for the required columns (assuming the third row contains the headers)
    header = {cell.value.strip() if cell.value else None: idx for idx, cell in enumerate(ws[3], 1)}
    print(f"Detected headers: {header}")  # Debugging: print detected headers

    # Ensure required headers are present
    required_headers = ['Start Date', 'End Date', 'Meeting Patterns', 'Course Listing']
    missing_headers = [col for col in required_headers if col not in header]
    if missing_headers:
        raise ValueError(f"Missing required headers: {missing_headers}")


    start_date_col = header.get('Start Date') - 1
    end_date_col = header.get('End Date') - 1
    meeting_patterns_col = header.get('Meeting Patterns') - 1
    class_name_col = header.get('Course Listing') - 1

    # Start processing from the third row (index 3)
    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip the first three rows
        start_date = row[start_date_col]
        end_date = row[end_date_col]
        meeting_patterns = row[meeting_patterns_col]
        class_name = row[class_name_col]

        # print(f"Row data: {row}")  # Print the row data for debugging
        if start_date and end_date and meeting_patterns and class_name:
            meeting_patterns = parse_course_entries(meeting_patterns)
            # Loop through the parsed meetings
            for meeting in meeting_patterns:
                day, time_interval, location = meeting
                start_time_str, end_time_str = time_interval.split(' - ')
                for meeting_date in get_week_dates(start_date, end_date, day):
                    start_datetime_str = f"{meeting_date.strftime('%Y-%m-%d')} {start_time_str}"
                    end_datetime_str = f"{meeting_date.strftime('%Y-%m-%d')} {end_time_str}"
                    start_datetime = est.localize(datetime.strptime(start_datetime_str, '%Y-%m-%d %I:%M %p'))
                    end_datetime = est.localize(datetime.strptime(end_datetime_str, '%Y-%m-%d %I:%M %p'))

                    event = Event()
                    event.name = class_name
                    event.begin = start_datetime
                    event.end = end_datetime
                    event.location = location
                    cal.events.add(event)
                    print(f"Class:{class_name} Start: {start_datetime}, End: {end_datetime}, Location: {location}")

                    # print(f"Class:{class_name} Day: {day}, Time: {time_interval}, Location: {location}")

            print("----------------------------")
        else:
            print(f"Skipping row with insufficient columns: {row}")

    with open(ics_path, 'w') as f:
        f.writelines(cal.serialize_iter())
    print("Finished printing class information.")


# Main function to handle command line arguments and execute the script
def main():
    if len(sys.argv) != 3:
        print("Usage: python script_name.py /path/to/View_My_Courses.xlsx /path/to/My_calendar.ics")
        sys.exit(1)

    file_path = sys.argv[1]
    ics_path = sys.argv[2]
    process_class_info(file_path, ics_path)


if __name__ == "__main__":
    main()